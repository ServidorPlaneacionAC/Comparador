import streamlit as st
import pandas as pd
import base64
from openpyxl.styles import PatternFill

# Tolerancia para la comparación de números decimales
TOLERANCIA_DECIMAL = 1e-9

# Función para encontrar filas con diferencias y marcar las celdas con un asterisco
def encontrar_filas_con_diferencias(df_base, df_comparar):
    # Identificar las filas que existen en el archivo a comparar pero no en la base de datos
    filas_nuevas = df_comparar.merge(df_base, how='outer', indicator=True).loc[lambda x: x['_merge'] == 'left_only'].drop(
        columns=['_merge'])

    # Obtener el DataFrame con filas que tienen diferencias
    df_diferencias = df_comparar[df_comparar.index.isin(filas_nuevas.index)].copy()

    # Marcar las celdas con un asterisco solo donde la información no es igual al archivo base
    for col in df_comparar.columns:
        # Comparar tipos de datos y manejar la igualdad numérica para tipos numéricos
        if pd.api.types.is_numeric_dtype(df_comparar[col]) and pd.api.types.is_numeric_dtype(df_base[col]):
            df_diferencias[col] = df_diferencias.apply(lambda x: f"{x[col]}*" if x.name in df_base.index and abs(x[col] - df_base.at[x.name, col]) > TOLERANCIA_DECIMAL else x[col], axis=1)
        else:
            df_diferencias[col] = df_diferencias.apply(lambda x: f"{x[col]}*" if x.name in df_base.index and x[col] != df_base.at[x.name, col] else x[col], axis=1)

    return df_diferencias

# Función para aplicar formato a las celdas con diferencias
def resaltar_diferencias(val):
    if '*' in str(val):
        return 'background-color: red'
    else:
        return ''

# Función para generar un enlace de descarga de un archivo binario
def get_binary_file_downloader_html(file_path, file_label='Archivo'):
    with open(file_path, 'rb') as f:
        data = f.read()
        base64_encoded = base64.b64encode(data).decode()
        return f'<a href="data:application/octet-stream;base64,{base64_encoded}" download="{file_path}">{file_label}</a>'

# Función para comparar los datos agrupados por material
def comparar_por_material(df_base, df_comparar):
    # Establecer la columna "Material" como índice
    df_base = df_base.set_index("Material")
    df_comparar = df_comparar.set_index("Material")

    # Agrupar por material y aplicar la función lambda para obtener el DataFrame agrupado
    df_base_grouped = df_base.groupby(level=0).apply(lambda x: x.droplevel(0)).reset_index(drop=True)
    df_comparar_grouped = df_comparar.groupby(level=0).apply(lambda x: x.droplevel(0)).reset_index(drop=True)

    # Unir los DataFrames en uno solo para detectar las diferencias
    df_merged = pd.merge(df_base_grouped, df_comparar_grouped, how='outer', on="Material", suffixes=('_base', '_comparar'))

    # Identificar las filas con diferencias en los datos
    df_diferencias = df_merged[df_merged.ne(df_merged.shift(axis=1)).any(axis=1)]

    return df_diferencias

# Titulo
st.title("Comparador de Datos Maestros")

# Para subir el archivo base en Excel
archivo_base = st.file_uploader("Cargar archivo base (base de datos)", type=["xlsx"])

# Para subir el archivo a comparar en Excel
archivo_comparar = st.file_uploader("Cargar archivo a comparar", type=["xlsx"])

if archivo_base and archivo_comparar:
    # Cargar los archivos
    df_base = pd.read_excel(archivo_base)
    df_comparar = pd.read_excel(archivo_comparar)

    # Verificar si los DataFrames son idénticos
    if df_base.equals(df_comparar):
        st.error("Los archivos son idénticos. No hay diferencias para resaltar.")
    elif df_base.columns.to_list() != df_comparar.columns.to_list():
        st.error("Los archivos no tienen las mismas columnas. Asegúrate de cargar archivos con las mismas columnas.")
    else:
        # Encontrar filas con diferencias y marcar celdas con un asterisco
        df_diferencias = encontrar_filas_con_diferencias(df_base, df_comparar)

        # Encontrar filas faltantes en comparación al archivo base
        df_faltantes = df_base.loc[~df_base.index.isin(df_comparar.index)]

        # Encontrar filas en el archivo a comparar que no están en el archivo base
        df_filas_en_comparar_no_en_base = df_comparar.loc[~df_comparar.index.isin(df_base.index)]

        # Mostrar el DataFrame con filas que tienen diferencias
        st.write("Informacion que tiene diferencias:")
        st.table(df_diferencias.style.applymap(resaltar_diferencias))

        # ... (código existente)

        # Botón para mostrar las filas en el archivo base correspondientes a las diferencias
        if st.button("Mostrar información del archivo base correspondiente a las diferencias"):
            st.write("Información del archivo base correspondiente a las diferencias:")
            # Filtrar el DataFrame base solo para las filas que tienen diferencias en el archivo a comparar
            df_base_diferencias = df_base[df_base.index.isin(df_diferencias.index)].copy()
            st.table(df_base_diferencias.style.applymap(resaltar_diferencias))

        # Botón para mostrar las filas en el archivo a comparar que no están en el archivo base
        if st.button("Mostrar informacion nueva en comparar "):
            st.write("Informacion en el archivo a comparar que no está en el archivo base:")
            st.table(df_filas_en_comparar_no_en_base.style.applymap(resaltar_diferencias))

        # Botón para mostrar la información faltante en comparación al archivo base
        if st.button("Mostrar información faltante en comparación al archivo base"):
            st.write("Información faltante en comparación al archivo base:")
            st.table(df_faltantes.style.applymap(resaltar_diferencias))

        # Botón para descargar la información en un archivo Excel con resaltado
if st.button("Descargar información en Excel con resaltado"):
    # Crear un objeto ExcelWriter para escribir en un solo archivo Excel
    with pd.ExcelWriter("informacion_comparada.xlsx", engine='openpyxl') as writer:
        # Escribir cada DataFrame en una pestaña diferente
        df_diferencias.to_excel(writer, sheet_name='Diferencias', index=True)
        df_filas_en_comparar_no_en_base.to_excel(writer, sheet_name='Filas_en_comparar_no_en_base', index=True)
        df_faltantes.to_excel(writer, sheet_name='Faltantes_en_base', index=True)

        # Obtener el objeto ExcelWriter y el objeto Workbook para aplicar el formato
        workbook = writer.book
        sheet = workbook['Diferencias']

        # Aplicar formato de resaltado a las celdas con diferencias
        for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column)):
            for cell in row:
                if '*' in str(cell.value):
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Enlace para descargar el archivo Excel
    st.markdown(
        get_binary_file_downloader_html("informacion_comparada.xlsx", 'Archivo Excel con resaltado'),
        unsafe_allow_html=True
    )
